import openpyxl
path= 'C:\\Users\\rling\\Py\\other\\files.xlsx'
WB = openpyxl.load_workbook(path)
sheet = WB.active

def FileNumCheck(x):
    for r in range(1, sheet.max_row):
        if str(x) == str(sheet.cell(row = r, column = 3).value):
            sheet.cell(row = r, column = 4).value = "basement"
            WB.save('C:\\Users\\rling\\Py\\other\\files.xlsx')
            return
        if r == (sheet.max_row)-1:
            print("could not find file number")
                  
userinput = 'word'
while userinput != "exit":
    userinput = str(input("input file number too add to the basement | input exit too close \n - "))
    FileNumCheck(userinput) 
   




                








